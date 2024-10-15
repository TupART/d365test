from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
import openpyxl
import os

app = Flask(__name__)

# Ruta para cargar archivos
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Página principal de subida de archivo
@app.route('/')
def upload_file():
    return render_template('upload.html')

# Procesar archivo y mostrar lista de checkboxes
@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        
        # Leer archivo .xlsx
        df = pd.read_excel(filepath)

        # Filtrar columnas "Name" y "Surname"
        names_surnames = df[['Name', 'Surname']].values.tolist()

        return render_template('upload.html', names_surnames=names_surnames, file_path=filepath)

# Procesar selección y completar PlantillaSTEP4.xlsx
@app.route('/submit', methods=['POST'])
def submit():
    selected_rows = request.form.getlist('rows')
    file_path = request.form['file_path']

    df = pd.read_excel(file_path)

    # Filtrar las filas seleccionadas
    selected_df = df.iloc[list(map(int, selected_rows))]

    # Cargar y rellenar PlantillaSTEP4.xlsx
    plantilla_path = 'uploads/PlantillaSTEP4.xlsx'
    plantilla = openpyxl.load_workbook(plantilla_path)
    ws = plantilla.active

    # Procesar y rellenar la plantilla con las filas seleccionadas
    for i, row in selected_df.iterrows():
        ws[f'C{i+7}'] = row['Name']  # Rellenar columna "Name"
        ws[f'D{i+7}'] = row['Surname']  # Rellenar columna "Surname"
        # Continuar rellenando las demás columnas con las reglas proporcionadas

    output_file = 'uploads/Completed_PlantillaSTEP4.xlsx'
    plantilla.save(output_file)

    return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
